import openpyxl

def get_excel_data(file_path, sheet_name):
    """
    Reads all rows from an Excel sheet (excluding header) and returns a list of tuples.
    Works for any number of columns.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    # Start from row 2 to skip header
    for row in range(2, sheet.max_row + 1):
        row_data = tuple(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1))
        data.append(row_data)

    return data