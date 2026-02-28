from openpyxl import load_workbook
import os

class ExcelUtils:

    def __init__(self):
        # Build dynamic path
        base_dir = os.path.dirname(os.path.dirname(__file__))
        file_path = os.path.join(base_dir, "data_for_test", "test_data.xlsx")

        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_test_data(self):
        data = []

        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            data.append({
                "first_name": row[0],
                "last_name": row[1],
                "postcode": row[2],
                "deposit": row[3],
                "withdraw": row[4]
            })

        return data